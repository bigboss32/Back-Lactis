"""Utilidades de exportación a Excel (openpyxl) y PDF (reportlab)."""
import io
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    HRFlowable,
    Image as RLImage,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

# Paleta de marca
BRAND = colors.HexColor("#1F4E78")
BRAND_LIGHT = colors.HexColor("#F2F6FA")
GOLD = colors.HexColor("#E39B1B")
GREY = colors.HexColor("#666666")
LOGO_PATH = Path(__file__).resolve().parent.parent / "assets" / "lactis-logo.png"

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Arial", size=10)
THIN_BORDER = Border(*[Side(style="thin")] * 4)


def _cell_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    return value


def rows_to_excel(
    *,
    title: str,
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
    sheet_name: str = "Datos",
    money_columns: Sequence[int] = (),
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 1))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(name="Arial", bold=True, size=13)
    title_cell.alignment = Alignment(horizontal="center")

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for r, row in enumerate(rows, start=4):
        for c, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=_cell_value(value))
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            if c in money_columns:
                cell.number_format = "$#,##0"

    for col in range(1, len(headers) + 1):
        max_len = max(
            [len(str(headers[col - 1]))] + [len(str(row[col - 1])) for row in rows if col <= len(row)] or [10]
        )
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_pdf(
    *,
    title: str,
    subtitle: str = "",
    sections: Sequence[dict[str, Any]],
) -> bytes:
    """Genera un PDF con secciones de tablas.

    Cada sección: {"heading": str, "headers": [...], "rows": [[...]], "col_widths": [...] opcional}
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=letter, topMargin=1.5 * cm, bottomMargin=1.5 * cm,
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitleQ", parent=styles["Title"], fontSize=15, spaceAfter=4)
    elements: list[Any] = [Paragraph(title, title_style)]
    if subtitle:
        elements.append(Paragraph(subtitle, styles["Normal"]))
    elements.append(Spacer(1, 10))

    for section in sections:
        if section.get("heading"):
            elements.append(Paragraph(section["heading"], styles["Heading3"]))
        data = [list(section["headers"])] + [
            [str(_cell_value(v)) if v is not None else "" for v in row] for row in section["rows"]
        ]
        table = Table(data, colWidths=section.get("col_widths"), repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F6FA")]),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]
            )
        )
        elements.append(table)
        elements.append(Spacer(1, 12))

    doc.build(elements)
    return buffer.getvalue()


def build_liquidacion_pdf(
    *,
    empresa_nombre: str,
    empresa_nit: str | None,
    empresa_ubicacion: str | None,
    folio: str,
    estado: str,
    emitido: str,
    tercero_label: str,
    tercero_nombre: str,
    tercero_detalle: str | None,
    periodo: str,
    detalle_headers: Sequence[str],
    detalle_rows: Sequence[Sequence[Any]],
    resumen_rows: Sequence[tuple[str, str, bool]],
    anticipos_rows: Sequence[Sequence[Any]] = (),
    observaciones: str | None = None,
) -> bytes:
    """Comprobante de liquidación con membrete, resumen, anticipos y firmas."""
    buffer = io.BytesIO()
    styles = getSampleStyleSheet()
    st_company = ParagraphStyle("Company", parent=styles["Title"], fontSize=16, textColor=BRAND, spaceAfter=0, leading=18, alignment=0)
    st_sub = ParagraphStyle("Sub", parent=styles["Normal"], fontSize=8, textColor=GREY, leading=11)
    st_doctitle = ParagraphStyle("DocT", parent=styles["Normal"], fontSize=12.5, textColor=BRAND, fontName="Helvetica-Bold", alignment=TA_RIGHT, leading=15)
    st_docmeta = ParagraphStyle("DocM", parent=styles["Normal"], fontSize=8.5, textColor=GREY, alignment=TA_RIGHT, leading=12)
    st_head = ParagraphStyle("Sec", parent=styles["Heading3"], fontSize=10.5, textColor=BRAND, spaceBefore=2, spaceAfter=4)
    st_lbl = ParagraphStyle("Lbl", parent=styles["Normal"], fontSize=7.5, textColor=GREY)
    st_val = ParagraphStyle("Val", parent=styles["Normal"], fontSize=9.5, fontName="Helvetica-Bold")
    st_obs = ParagraphStyle("Obs", parent=styles["Normal"], fontSize=9, leading=13)
    st_sign = ParagraphStyle("Sign", parent=styles["Normal"], fontSize=8.5, alignment=TA_CENTER, textColor=GREY, leading=12)

    # --- Encabezado: logo + empresa + datos del comprobante
    company_block: list[Any] = [Paragraph(empresa_nombre, st_company)]
    sub = " · ".join(
        p for p in [f"NIT {empresa_nit}" if empresa_nit else None, empresa_ubicacion] if p
    )
    if sub:
        company_block.append(Paragraph(sub, st_sub))
    doc_block = [
        Paragraph("COMPROBANTE DE LIQUIDACIÓN", st_doctitle),
        Paragraph(f"N.º {folio}", st_docmeta),
        Paragraph(f"Estado: <b>{estado.upper()}</b>", st_docmeta),
        Paragraph(f"Emitido: {emitido}", st_docmeta),
    ]
    logo_cell: Any = (
        RLImage(str(LOGO_PATH), width=1.4 * cm, height=1.4 * cm) if LOGO_PATH.exists() else ""
    )
    header = Table([[logo_cell, company_block, doc_block]], colWidths=[1.7 * cm, 8.6 * cm, 7.0 * cm])
    header.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (0, 0), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    elements: list[Any] = [
        header,
        HRFlowable(width="100%", thickness=1.2, color=BRAND, spaceBefore=6, spaceAfter=10),
    ]

    # --- Datos del tercero
    info_rows = [
        [Paragraph(tercero_label, st_lbl), Paragraph(tercero_nombre, st_val),
         Paragraph("Período", st_lbl), Paragraph(periodo, st_val)],
    ]
    if tercero_detalle:
        info_rows.append(
            [Paragraph("Ruta / vereda", st_lbl), Paragraph(tercero_detalle, st_val),
             Paragraph("Comprobante", st_lbl), Paragraph(f"N.º {folio}", st_val)]
        )
    info = Table(info_rows, colWidths=[2.6 * cm, 6.2 * cm, 2.6 * cm, 5.9 * cm])
    info.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), BRAND_LIGHT),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#D6E0EA")),
                ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("LEFTPADDING", (0, 0), (-1, -1), 8), ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    elements += [info, Spacer(1, 12)]

    # --- Detalle diario
    elements.append(Paragraph("Detalle diario", st_head))
    det_data = [list(detalle_headers)] + [
        [str(_cell_value(v)) if v is not None else "" for v in row] for row in detalle_rows
    ]
    det = Table(det_data, repeatRows=1, hAlign="LEFT")
    det.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), BRAND),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D6E0EA")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, BRAND_LIGHT]),
                ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    elements += [det, Spacer(1, 12)]

    # --- Resumen (con VALOR TOTAL y SALDO destacados)
    elements.append(Paragraph("Resumen de liquidación", st_head))
    res = Table([[c, v] for (c, v, _) in resumen_rows], colWidths=[6 * cm, 5 * cm], hAlign="RIGHT")
    res_style = [
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("LINEBELOW", (0, 0), (-1, -1), 0.4, colors.HexColor("#E6E6E6")),
        ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 8), ("RIGHTPADDING", (0, 0), (-1, -1), 8),
    ]
    for i, (_, _, resaltado) in enumerate(resumen_rows):
        if resaltado:
            res_style += [
                ("BACKGROUND", (0, i), (-1, i), BRAND_LIGHT),
                ("FONTNAME", (0, i), (-1, i), "Helvetica-Bold"),
                ("TEXTCOLOR", (0, i), (-1, i), BRAND),
                ("FONTSIZE", (0, i), (-1, i), 10),
            ]
    res.setStyle(TableStyle(res_style))
    elements += [res, Spacer(1, 12)]

    # --- Anticipos aplicados
    if anticipos_rows:
        elements.append(Paragraph("Anticipos aplicados", st_head))
        ant_data = [["Fecha", "Valor", "Observaciones"]] + [
            [str(_cell_value(v)) if v is not None else "" for v in row] for row in anticipos_rows
        ]
        ant = Table(ant_data, colWidths=[3 * cm, 3 * cm, 10.9 * cm], repeatRows=1, hAlign="LEFT")
        ant.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), BRAND),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D6E0EA")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, BRAND_LIGHT]),
                    ("ALIGN", (1, 1), (1, -1), "RIGHT"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ]
            )
        )
        elements += [ant, Spacer(1, 12)]

    # --- Observaciones
    if observaciones:
        elements.append(Paragraph("Observaciones", st_head))
        box = Table([[Paragraph(observaciones, st_obs)]], colWidths=[16.9 * cm])
        box.setStyle(
            TableStyle(
                [
                    ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#D6E0EA")),
                    ("BACKGROUND", (0, 0), (-1, -1), colors.white),
                    ("TOPPADDING", (0, 0), (-1, -1), 8), ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                    ("LEFTPADDING", (0, 0), (-1, -1), 8), ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ]
            )
        )
        elements += [box]

    # --- Firmas
    elements.append(Spacer(1, 28))
    firma = Table(
        [
            ["", ""],
            [
                Paragraph(f"Entregué conforme<br/>{empresa_nombre}", st_sign),
                Paragraph(f"Recibí conforme<br/>{tercero_nombre}", st_sign),
            ],
        ],
        colWidths=[8.4 * cm, 8.4 * cm],
        rowHeights=[0.9 * cm, None],
    )
    firma.setStyle(
        TableStyle(
            [
                ("LINEABOVE", (0, 1), (0, 1), 0.6, colors.black),
                ("LINEABOVE", (1, 1), (1, 1), 0.6, colors.black),
                ("TOPPADDING", (0, 1), (-1, 1), 4),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("LEFTPADDING", (0, 0), (-1, -1), 20), ("RIGHTPADDING", (0, 0), (-1, -1), 20),
            ]
        )
    )
    elements.append(firma)

    def _footer(canvas: Any, doc_: Any) -> None:
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor("#D6E0EA"))
        canvas.setLineWidth(0.5)
        canvas.line(1.5 * cm, 1.3 * cm, letter[0] - 1.5 * cm, 1.3 * cm)
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(GREY)
        canvas.drawString(1.5 * cm, 1.0 * cm, f"Generado por Lactis · {emitido}")
        canvas.drawRightString(letter[0] - 1.5 * cm, 1.0 * cm, f"Página {doc_.page}")
        canvas.restoreState()

    doc = SimpleDocTemplate(
        buffer, pagesize=letter, topMargin=1.4 * cm, bottomMargin=1.8 * cm,
        leftMargin=1.5 * cm, rightMargin=1.5 * cm, title=f"Liquidación {folio}",
    )
    doc.build(elements, onFirstPage=_footer, onLaterPages=_footer)
    return buffer.getvalue()
